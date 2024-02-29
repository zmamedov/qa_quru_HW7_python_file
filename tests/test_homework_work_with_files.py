from zipfile import ZipFile
from tests.settings import ARCHIVE
from pypdf import PdfReader
from openpyxl import load_workbook
import csv


def test_pdf_in_archive():
    with ZipFile(ARCHIVE) as archive:
        with archive.open('file_pdf.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            number_of_pages = len(reader.pages)
            page = reader.pages[1].extract_text()

            assert number_of_pages == 46
            assert 'Agile Testing' in page


def test_xlsx_in_archive():
    with ZipFile(ARCHIVE) as archive:
        with archive.open('file_xlsx.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            number_of_sheets = len(workbook.sheetnames)
            sheet = workbook.active
            cell_1 = sheet.cell(row=1, column=1).value

            assert number_of_sheets == 15
            assert cell_1 == 'Дата'


def test_csv_in_archive():
    with ZipFile(ARCHIVE) as archive:
        with archive.open('file_csv.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csv_reader = list((csv.reader(content.splitlines())))
            number_of_elements = len(csv_reader)
            second_row = csv_reader[1]

            assert number_of_elements == 1001
            assert second_row[0] == '1'
